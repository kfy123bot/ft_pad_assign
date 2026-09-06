#!/usr/bin/env python3
"""
gen_ug_excel.py — Generate ft_pad_assign User Guide Excel file.

Output: docs/ft_pad_assign_ug.xlsx  (5 sheets)

Sheet 1: 使用說明       — Tool overview, CLI flags, output files
Sheet 2: Header 欄位    — Header key reference table
Sheet 3: 資料欄位       — 11 data column reference with aliases
Sheet 4: 特殊列類型     — NC / DOWNBOND / POWERCUT / Inner Bond matrix
Sheet 5: Pin List 模板   — Ready-to-fill template (save as CSV to use directly)
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
C_TITLE     = "1F4E79"   # dark navy — sheet titles
C_HEADER    = "2E75B6"   # blue — table headers
C_HEADER_FG = "FFFFFF"   # white text on header
C_REQUIRED  = "FCE4D6"   # light orange — required fields
C_OPTIONAL  = "EBF3FB"   # light blue — optional fields
C_SPECIAL   = "FFF2CC"   # light yellow — special/note cells
C_NC        = "D9D9D9"   # grey — NC row
C_DOWNBOND  = "BDD7EE"   # light blue — DOWNBOND row
C_POWERCUT  = "F4B8D2"   # light pink — POWERCUT row
C_INNERBOND = "FCE4D6"   # light orange — Inner Bond row
C_TEMPLATE  = "E2EFDA"   # light green — template header bg
C_NOTE      = "FFFF99"   # yellow — note / guide column
C_ALT       = "F2F2F2"   # light grey — alternating rows
C_PASS      = "C6EFCE"   # green — pass/ok
C_WARN      = "FFEB9C"   # amber — warning

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, fg=None, bg=None,
             h_align="left", wrap=False, italic=False, font_size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold, color=fg or "000000", size=font_size, italic=italic)
    if bg:
        cell.fill = fill(bg)
    cell.alignment = align(h=h_align, wrap=wrap)
    cell.border = thin_border()
    return cell

def header_row(ws, row, cols, bg=C_HEADER):
    for col, val in enumerate(cols, 1):
        set_cell(ws, row, col, val, bold=True, fg=C_HEADER_FG, bg=bg, h_align="center")

def title_row(ws, row, text, ncols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=C_HEADER_FG, size=13)
    c.fill = fill(C_TITLE)
    c.alignment = align(h="center")

def section_row(ws, row, text, ncols=10, bg="305496"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = fill(bg)
    c.alignment = align(h="left")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------------------
# Sheet 1 — 使用說明
# ---------------------------------------------------------------------------
def build_sheet1(wb):
    ws = wb.active
    ws.title = "使用說明"
    ws.sheet_view.showGridLines = True
    set_col_widths(ws, [18, 14, 50, 30, 20, 20, 20, 20, 20, 20])

    r = 1
    title_row(ws, r, "FT_PAD_ASSIGN 使用說明書  |  ft_pad_assign.py v5", ncols=6)
    r += 1
    set_cell(ws, r, 1, "讀取 Pin List CSV，產生封裝示意圖（PDF）、補齊 Pin List 及 APR I/O 約束檔",
             italic=True, h_align="left")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2

    # --- CLI section ---
    section_row(ws, r, "  命令列用法", ncols=6)
    r += 1
    set_cell(ws, r, 1, "python3 bin/ft_pad_assign.py  -list <pin_list_file>  [選項]",
             bold=True, bg=C_SPECIAL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    header_row(ws, r, ["選項", "類型", "說明", "範例"])
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

    cli_data = [
        ("-list <file>", "必要", "Pin List 輸入檔（.csv 或 .pin_list）",
         "examples/qfn48.8028.pin_list.csv"),
        ("-o <目錄>",    "選填", "輸出目錄（預設：當前目錄）",
         "-o output"),
        ("-all",         "選填", "產生所有輸出（等同同時指定以下所有旗標）",
         "-all"),
        ("-apr",         "選填", "產生 APR PDF（Die Pad 佈局圖）",
         "-apr"),
        ("-pkg",         "選填", "產生 PKG PDF（封裝腳位示意圖）",
         "-pkg"),
        ("-combined",    "選填", "產生 Combined PDF（PKG+APR 合併圖 + Bonding Wire）",
         "-combined"),
        ("-c",           "選填", "產生 .new、.new.csv；搭配 -v 還產生 .inn.const / .icc2.const",
         "-c"),
        ("-stagger",     "選填", "執行 I/O Stagger 密度檢查，產生 _stagger.rpt",
         "-stagger"),
        ("-stagger-max N","選填","連續 I/O pin 警告閾值（預設：8）",
         "-stagger-max 6"),
        ("-v [files]",   "選填", "Verilog 網表（自動填入 IO_CELL_NAME、DIRECTION）",
         "-v top.v"),
    ]
    for row_data in cli_data:
        r += 1
        bg = C_REQUIRED if row_data[1] == "必要" else None
        for col, val in enumerate(row_data[:3], 1):
            set_cell(ws, r, col, val, bg=bg, wrap=(col == 3))
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        set_cell(ws, r, 4, row_data[3], italic=True, bg=bg)

    r += 2
    section_row(ws, r, "  輸出檔案說明", ncols=6)
    r += 1
    header_row(ws, r, ["檔案", "產生條件", "說明", "", "", ""])
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

    output_data = [
        ("<proj>.log",          "始終",                "執行日誌（INFO / WARN / ERROR / FATAL）"),
        ("<proj>.new",          "-c 或 -all",          "補齊後的 Pin List（固定寬度文字格式）"),
        ("<proj>.new.csv",      "-c 或 -all",          "補齊後的 Pin List（CSV 格式，可再次作為輸入）"),
        ("<proj>_pkg.pdf",      "-pkg 或 -all",        "封裝腳位示意圖（外觀視角）"),
        ("<proj>_apr.pdf",      "-apr 或 -all",        "Die Pad 佈局圖（晶片內部視角）"),
        ("<proj>_combined.pdf", "-combined 或 -all",   "PKG + APR 合併圖 + Bonding Wire"),
        ("<proj>_chip.inn.const", "-c + -v + -all",   "Innovus I/O 擺放約束（Tcl 格式）"),
        ("<proj>_chip.icc2.const","-c + -v + -all",   "ICC2 I/O 擺放約束（Tcl 命令格式）"),
        ("<proj>_stagger.rpt",  "-stagger 或 -all",    "I/O Stagger 密度警告報告"),
    ]
    for row_data in output_data:
        r += 1
        set_cell(ws, r, 1, row_data[0], bold=True)
        set_cell(ws, r, 2, row_data[1], h_align="center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        set_cell(ws, r, 3, row_data[2], wrap=True)

    r += 2
    section_row(ws, r, "  安裝需求", ncols=6)
    r += 1
    header_row(ws, r, ["套件", "必要/選填", "安裝指令", "", "", ""])
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    r += 1
    set_cell(ws, r, 1, "Python 3.7+"); set_cell(ws, r, 2, "必要")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    set_cell(ws, r, 3, "（標準 Python 安裝）", italic=True)
    r += 1
    set_cell(ws, r, 1, "reportlab", bold=True)
    set_cell(ws, r, 2, "必要", bg=C_REQUIRED)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    set_cell(ws, r, 3, "pip3 install reportlab", italic=True)
    r += 1
    set_cell(ws, r, 1, "openpyxl")
    set_cell(ws, r, 2, "選填（Excel 模板）")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    set_cell(ws, r, 3, "pip3 install openpyxl", italic=True)

    ws.row_dimensions[1].height = 22
    return ws

# ---------------------------------------------------------------------------
# Sheet 2 — Header 欄位
# ---------------------------------------------------------------------------
def build_sheet2(wb):
    ws = wb.create_sheet("Header 欄位")
    set_col_widths(ws, [22, 12, 28, 30, 15, 40])

    r = 1
    title_row(ws, r, "Pin List Header 欄位說明", ncols=6)
    r += 2

    section_row(ws, r, "  必要 Header（缺少時工具無法正常運作）", ncols=6)
    r += 1
    header_row(ws, r, ["Key", "必要/選填", "格式", "範例", "別名", "說明"])
    r += 1

    req_data = [
        ("PRODUCTION NO", "必要", "任意字串",
         "PRJ8028_QFN48_TEST", "PROJECT NO",
         "專案代號。決定所有輸出檔名前綴。非字母數字字元自動轉為底線 _。"),
        ("PACKAGE", "必要", "<類型> <L> <B> <R> <T>",
         "48QFN 12 12 12 12", "—",
         "封裝類型及四邊 pin 數量（L=左, B=下, R=右, T=上）。工具依此驗證 pin 數並計算 PKG_LOC。"),
        ("VERSION", "必要", "任意字串",
         "V1.0_20240418", "—",
         "版本號，顯示於 PDF 頁首。"),
    ]
    for d in req_data:
        set_cell(ws, r, 1, d[0], bold=True, bg=C_REQUIRED)
        set_cell(ws, r, 2, d[1], h_align="center", bg=C_REQUIRED)
        set_cell(ws, r, 3, d[2], bg=C_REQUIRED)
        set_cell(ws, r, 4, d[3], italic=True, bg=C_REQUIRED)
        set_cell(ws, r, 5, d[4], bg=C_REQUIRED)
        set_cell(ws, r, 6, d[5], wrap=True, bg=C_REQUIRED)
        ws.row_dimensions[r].height = 40
        r += 1

    r += 1
    section_row(ws, r, "  選填 Header", ncols=6)
    r += 1
    header_row(ws, r, ["Key", "必要/選填", "格式", "範例", "預設值", "說明"])
    r += 1

    opt_data = [
        ("PKG_TOP_LEFT_PIN", "選填", "正整數",
         "15",  "1（不 shift）",
         "指定 L 邊第一根 pin 的原始編號。不為 1 時觸發 Ring Shift：\n"
         "資料循環重排 → PKG_NUM 從 1 重編 → 重置為 1。"),
        ("DIE_SIZE", "選填", "<寬um>x<高um>",
         "2414x1415", "—（正方形）",
         "Die 實際尺寸（微米）。填入後 APR PDF 依比例繪製矩形 die；\n"
         "PKG PDF 依 PKG body / die 比例縮放框架。"),
        ("PKG_SIZE", "選填", "<寬um>x<高um>",
         "7000x7000", "查封裝尺寸表",
         "Package body 尺寸（微米）。填入後 PKG PDF 使用此值；\n"
         "省略時查內建 QFN_BODY_SIZES 表（參見「QFN 封裝尺寸」sheet）。"),
    ]
    for d in opt_data:
        set_cell(ws, r, 1, d[0], bold=True, bg=C_OPTIONAL)
        set_cell(ws, r, 2, d[1], h_align="center", bg=C_OPTIONAL)
        set_cell(ws, r, 3, d[2], bg=C_OPTIONAL)
        set_cell(ws, r, 4, d[3], italic=True, bg=C_OPTIONAL)
        set_cell(ws, r, 5, d[4], bg=C_OPTIONAL)
        set_cell(ws, r, 6, d[5], wrap=True, bg=C_OPTIONAL)
        ws.row_dimensions[r].height = 52
        r += 1

    r += 2
    section_row(ws, r, "  書寫規則", ncols=6)
    r += 1
    notes = [
        "• Key 不區分大小寫：production no、PRODUCTION NO、Production No 均可",
        "• CSV 格式中行末可有多餘逗號（如 PRODUCTION NO : PRJ8028,,,,,,,），自動忽略",
        "• Header 行識別條件：行內含 ':'，且不以數字、'(' 或 'D1' 開頭",
        "• Header 行順序不固定，不認識的 Key 一律忽略（不報錯）",
        "• 輸出的 .new 和 .new.csv 統一將 PRODUCTION NO 正規化為 PROJECT NO",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        set_cell(ws, r, 1, note, wrap=True)
        ws.row_dimensions[r].height = 20
        r += 1

    return ws

# ---------------------------------------------------------------------------
# Sheet 3 — 資料欄位
# ---------------------------------------------------------------------------
def build_sheet3(wb):
    ws = wb.create_sheet("資料欄位")
    set_col_widths(ws, [18, 24, 20, 32, 36, 20])

    r = 1
    title_row(ws, r, "Pin List 資料欄位說明（11 個欄位）", ncols=6)
    r += 1
    set_cell(ws, r, 1, "CSV 表頭行標準寫法：", bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    set_cell(ws, r, 2,
             "PKG_NUM, PKG_PIN_NAME, DIE_NUM, DIE_PIN_NAME, IO_CELL_NAME, "
             "PKG_LOC, DIE_LOC, DIRECTION, LOAD, SLEW, SSO",
             bold=True, italic=True)
    r += 2

    header_row(ws, r, ["欄位名稱", "別名", "允許值", "特殊值說明", "行為說明", "必/選填"])
    r += 1

    fields = [
        # (name, alias, allowed, special, behavior, req)
        ("PKG_NUM", "PIN_NUM",
         "正整數 / 0 / - / D1.xx / (D1.xx) / Inner_bond",
         "0=共用/無pkg pin，D1.xx=Inner Bond源端，(D1.xx)=目標端，Inner_bond=模板列",
         "正整數→正常pkg pin；0→跳過PKG PDF；D1.xx→繪製Inner Bond線條。Ring Shift後正整數重編為1,2,3...",
         "必要", C_REQUIRED),
        ("PKG_PIN_NAME", "PACKAGE_PIN\nPKG_PIN",
         "任意字串 / -",
         "—",
         "PKG PDF 上顯示的 pin 標籤。空或 - 時 fallback 到 DIE_PIN_NAME。",
         "選填", None),
        ("DIE_NUM", "DIE_PAD_NUM",
         "正整數 / 0 / -",
         "0=無die pad (NC/DOWNBOND)，-=無效列",
         "工具從L邊第一個非NC pin起重編為1,2,3...。相同原始DIE_NUM的列共用新編號（APR PDF去重）。",
         "必要", C_REQUIRED),
        ("DIE_PIN_NAME", "PIN_NAME",
         "任意字串 / NC / DOWNBOND / *POWERCUT* / -",
         "NC=黑色方塊，DOWNBOND=藍色方塊+接地符號，POWERCUT=APR黑色方塊",
         "決定pin類型及PDF顯示行為（最關鍵欄位）。含%時分段：PKG顯示第一段，APR顯示最後段。",
         "必要", C_REQUIRED),
        ("IO_CELL_NAME", "CELL_NAME\nIO_CELL\nIOCELL",
         "hierarchical inst name / NOT_FOUND / -",
         "—",
         "I/O Cell 實例名稱。搭配 -v 時，空或NOT_FOUND自動從Verilog查找。用於產生.inn.const/.icc2.const。",
         "選填", None),
        ("PKG_LOC", "LOCATION\nPIN_LOCA",
         "L / B / R / T / -",
         "—",
         "工具根據PKG_NUM自動計算並覆蓋此欄位。填 - 即可（不需手動填）。分配邏輯見PACKAGE header說明。",
         "自動", C_SPECIAL),
        ("DIE_LOC", "DIE_PAD_NUM_LOC\nDIE_LOCA",
         "L / B / R / T / -",
         "—",
         "PKG_TOP_LEFT_PIN≠1時工具自動跟隨PKG_LOC。PKG_TOP_LEFT_PIN=1時需手動填寫。",
         "半自動", C_OPTIONAL),
        ("DIRECTION", "IO_DIRECTION\nIO_TYPE\nDIR",
         "P / G / I / O / B / -",
         "P=電源(紅色wire)，G=接地(藍色wire)，I/O/B=訊號(灰色wire)",
         "影響Combined PDF wire顏色。搭配-v時，-的訊號pin自動從Verilog port方向填入。",
         "選填", None),
        ("LOAD", "CAP\nCAPACITANCE",
         "數值 / -",
         "—",
         "電容值。用於Stagger報告。不影響PDF顯示。",
         "選填", None),
        ("SLEW", "TRANSITION\nSLEW_RATE",
         "數值 / -",
         "—",
         "轉換率。用於Stagger報告。不影響PDF顯示。",
         "選填", None),
        ("SSO", "SSO_RATIO",
         "任意字串 / -",
         "RX, TX, LDO, AUDIO... 等功能群組標籤",
         "用於Stagger報告及視覺分組。不影響PDF顯示。",
         "選填", None),
    ]

    for i, f in enumerate(fields):
        bg = f[6] if f[6] else (C_ALT if i % 2 == 0 else None)
        set_cell(ws, r, 1, f[0], bold=True, bg=bg)
        set_cell(ws, r, 2, f[1], bg=bg, wrap=True)
        set_cell(ws, r, 3, f[2], bg=bg, wrap=True)
        set_cell(ws, r, 4, f[3], bg=bg, wrap=True)
        set_cell(ws, r, 5, f[4], bg=bg, wrap=True)
        set_cell(ws, r, 6, f[5], h_align="center", bg=bg)
        ws.row_dimensions[r].height = 52
        r += 1

    r += 2
    section_row(ws, r, "  DIE_PIN_NAME 的 % 分隔規則", ncols=6)
    r += 1
    set_cell(ws, r, 1, "格式：<PKG顯示名>%<類型>%<APR顯示名>", bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    header_row(ws, r, ["輸入值", "PKG PDF 標籤", "APR PDF 標籤", "", "", ""])
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    r += 1
    pct_data = [
        ("SCL", "SCL", "SCL"),
        ("VDD11%C%U_VDD11_APR5", "VDD11", "U_VDD11_APR5"),
        ("VDD33_IOB%IO%U_AIP_TOP/U_VDD33_IOB0", "VDD33_IOB", "U_VDD33_IOB0"),
    ]
    for d in pct_data:
        set_cell(ws, r, 1, d[0], italic=True)
        set_cell(ws, r, 2, d[1], bold=True, bg=C_PASS)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        set_cell(ws, r, 3, d[2], bold=True, bg=C_OPTIONAL)
        r += 1

    return ws

# ---------------------------------------------------------------------------
# Sheet 4 — 特殊列類型
# ---------------------------------------------------------------------------
def build_sheet4(wb):
    ws = wb.create_sheet("特殊列類型")
    set_col_widths(ws, [16, 12, 12, 12, 14, 14, 14, 36])

    r = 1
    title_row(ws, r, "特殊列類型說明", ncols=8)
    r += 2

    # --- Summary matrix ---
    header_row(ws, r,
               ["類型", "PKG_NUM", "DIE_NUM", "DIE_PIN_NAME",
                "PKG PDF", "APR PDF", "Combined PDF", "說明"],
               bg=C_HEADER)
    r += 1

    types = [
        # type, pkg_num, die_num, die_pin_name, pkg_pdf, apr_pdf, combined, note, bg
        ("NC", "正整數", "0", "NC",
         "黑色方塊", "跳過", "跳過",
         "No Connect。封裝腳位存在但無 die pad 連接。", C_NC),
        ("DOWNBOND", "正整數", "0", "DOWNBOND",
         "藍色方塊", "跳過", "接地符號（倒T）",
         "封裝腳位直接接地。Combined PDF 繪製藍色倒 T 接地符號。DIRECTION 需填 G。", C_DOWNBOND),
        ("POWERCUT", "0", "正整數", "含 POWERCUT",
         "跳過", "黑色方塊", "跳過",
         "Die 電源切割點，無對應封裝腳位。", C_POWERCUT),
        ("共用 Die Pad", "0", "正整數", "一般名稱",
         "跳過", "顯示（去重）", "繪製 N 條 wire",
         "同一 die pad 出現 N 次（PKG_NUM=0），Combined PDF 繪製 N 條 wire。", None),
        ("Inner Bond\n（無括號）", "D1.xx", "目標 DIE_NUM", "一般名稱",
         "延伸線", "延伸線", "紅色 wire",
         "D1.77 + DIE_NUM=42 → 從 pad77 畫線到 pad42。\n對稱時為實線，單向時為虛線+ERROR。", C_INNERBOND),
        ("Inner Bond\n（有括號）", "(D1.xx)", "目標 DIE_NUM", "一般名稱",
         "延伸線", "延伸線", "紅色 wire（反向）",
         "(D1.77) + DIE_NUM=42 → 從 pad42 畫線到 pad77（括號=反向）。", C_INNERBOND),
        ("Inner_bond 模板", "Inner_bond", "序號", "任意（佔位）",
         "跳過", "跳過", "跳過",
         "位於檔案尾部的備注列。不影響 PDF 和 pin 計數。原樣保留到 .new 輸出。", C_SPECIAL),
        ("佔位空列", "0 或 -", "0 或 -", "-",
         "跳過", "跳過", "跳過",
         "PKG_NUM 和 DIE_NUM 同為空/0/-，無實際意義。.new 輸出時跳過。", C_ALT),
    ]

    for t in types:
        bg = t[8]
        for col, val in enumerate(t[:8], 1):
            set_cell(ws, r, col, val, bg=bg, wrap=(col == 8),
                     h_align="center" if col in (2, 3, 5, 6, 7) else "left")
        ws.row_dimensions[r].height = 48
        r += 1

    r += 2
    section_row(ws, r, "  Inner Bond 方向對照表", ncols=8)
    r += 1
    header_row(ws, r, ["PKG_NUM", "DIE_NUM", "連線方向", "線條樣式", "", "", "", ""])
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    r += 1
    ib_examples = [
        ("D1.77", "42", "pad77 → pad42（無括號：從xx到目標）", "A→B + B→A 同時存在 → 實線（solid）"),
        ("(D1.77)", "42", "pad42 → pad77（有括號：反向）", "只有單方向 → 虛線（dashed） + ERROR"),
    ]
    for e in ib_examples:
        set_cell(ws, r, 1, e[0], bg=C_INNERBOND, bold=True)
        set_cell(ws, r, 2, e[1], bg=C_INNERBOND, h_align="center")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        set_cell(ws, r, 3, e[2], bg=C_INNERBOND, wrap=True)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
        set_cell(ws, r, 5, e[3], bg=C_INNERBOND, wrap=True)
        ws.row_dimensions[r].height = 36
        r += 1

    return ws

# ---------------------------------------------------------------------------
# Sheet 5 — Pin List 模板
# ---------------------------------------------------------------------------
def build_sheet5(wb):
    ws = wb.create_sheet("Pin List 模板")
    set_col_widths(ws, [22, 18, 12, 26, 22, 10, 10, 12, 8, 8, 10, 4, 28, 28, 28, 28])

    r = 1
    title_row(ws, r, "Pin List 模板 — 可另存為 .csv 後直接作為輸入使用", ncols=16)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
    c = ws.cell(row=r, column=1,
                value="【說明】第 1～11 欄為 pin list 欄位；第 12 欄為分隔線（可刪）；"
                      "第 13～16 欄為填寫說明（儲存為 CSV 後成為多餘尾部逗號，工具自動忽略）。")
    c.font = Font(color="595959", size=9, italic=True)
    c.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 2

    def note(text):
        return text

    def header_section_row(ws, r, key, value, note_text, bg=C_TEMPLATE):
        set_cell(ws, r, 1, f"{key} : {value}", bold=True, bg=bg)
        for col in range(2, 12):
            ws.cell(row=r, column=col).fill = fill(bg)
            ws.cell(row=r, column=col).border = thin_border()
        set_cell(ws, r, 12, "", bg=bg)
        ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=16)
        set_cell(ws, r, 13, note_text, bg=C_NOTE, italic=True, wrap=True)
        ws.row_dimensions[r].height = 20

    # --- Header section ---
    section_row(ws, r, "  ▼ 區塊一：Header 元資料（每行格式：KEY : VALUE）", ncols=16, bg="5B9BD5")
    r += 1
    header_section_row(ws, r, "PRODUCTION NO", "YOUR_PROJECT_NAME",
                       "★必要。專案代號 → 決定輸出檔名前綴。非字母數字自動轉底線。")
    r += 1
    header_section_row(ws, r, "PKG_TOP_LEFT_PIN", "1",
                       "選填。L邊第一根pin的原始編號。預設=1（不shift）。非1時觸發Ring Shift重排。")
    r += 1
    header_section_row(ws, r, "PACKAGE", "48QFN 12 12 12 12",
                       "★必要。格式：<類型> <L數> <B數> <R數> <T數>。工具用此驗證pin數並計算PKG_LOC。")
    r += 1
    header_section_row(ws, r, "VERSION", "V1.0_YYYYMMDD",
                       "必要。版本號，顯示於PDF頁首。")
    r += 1
    header_section_row(ws, r, "DIE_SIZE", "WIDTHxHEIGHT",
                       "選填。Die尺寸（微米），如 2414x1415。填入後PDF依比例繪製矩形die。")
    r += 1
    header_section_row(ws, r, "PKG_SIZE", "WIDTHxHEIGHT",
                       "選填。Package body尺寸（微米），如 6000x6000。省略時查內建封裝尺寸表。")
    r += 2

    # --- Data table header ---
    section_row(ws, r, "  ▼ 區塊二：資料表（空白行後開始）", ncols=16, bg="5B9BD5")
    r += 1

    # Column header row for data table
    col_names = ["PKG_NUM", "PKG_PIN_NAME", "DIE_NUM", "DIE_PIN_NAME",
                 "IO_CELL_NAME", "PKG_LOC", "DIE_LOC", "DIRECTION",
                 "LOAD", "SLEW", "SSO"]
    for col, name in enumerate(col_names, 1):
        set_cell(ws, r, col, name, bold=True, fg=C_HEADER_FG, bg=C_HEADER, h_align="center")
    set_cell(ws, r, 12, "", bg=C_HEADER)
    note_headers = ["欄位說明", "允許值", "重要規則", "特殊值"]
    for col, nh in enumerate(note_headers, 13):
        set_cell(ws, r, col, nh, bold=True, bg="808000", fg="FFFFFF", h_align="center")
    r += 1

    # Sub-header: field descriptions
    field_descs = [
        "封裝腳位編號", "封裝側pin名稱", "Die墊片編號", "Die側pin名稱",
        "IO Cell名稱", "封裝邊", "Die邊", "方向", "電容", "轉換率", "SSO"
    ]
    field_vals = [
        "正整數/0/D1.xx/(D1.xx)/Inner_bond/-",
        "字串/-",
        "正整數/0/-",
        "字串/NC/DOWNBOND/*POWERCUT*/-",
        "hierarchical inst / -",
        "L/B/R/T/-",
        "L/B/R/T/-",
        "P/G/I/O/B/-",
        "數值/-",
        "數值/-",
        "字串/-"
    ]
    field_rules = [
        "工具重編為1,2,3...；0=共用pad；Ring Shift後重編",
        "空時fallback到DIE_PIN_NAME",
        "工具從L邊起重編；相同原始值共用新號",
        "決定pin類型！含%時分段顯示：PKG=第1段，APR=最後段",
        "搭配-v自動填入；用於.inn.const/.icc2.const",
        "工具自動計算，手動值被覆蓋。填-即可",
        "Ring Shift後自動跟PKG_LOC；否則需手動填",
        "影響wire顏色；搭配-v自動填入I/O方向",
        "用於Stagger報告，不影響PDF",
        "用於Stagger報告，不影響PDF",
        "用於Stagger報告，不影響PDF"
    ]
    field_special = [
        "D1.xx=Inner Bond無括號；(D1.xx)=有括號反向",
        "（無特殊值）",
        "0=NC/DOWNBOND或共用pad",
        "NC=黑方塊；DOWNBOND=藍方塊；POWERCUT=APR黑方塊",
        "NOT_FOUND=等待-v補全",
        "（自動填入）",
        "（部分自動填入）",
        "P=紅wire；G=藍wire；I/O/B=灰wire",
        "（無特殊值）",
        "（無特殊值）",
        "常用：RX TX LDO AUDIO PLL"
    ]
    for col, d in enumerate(field_descs, 1):
        set_cell(ws, r, col, d, bold=False, bg=C_ALT, h_align="center", font_size=9)
    set_cell(ws, r, 12, "", bg=C_ALT)
    for col, v in enumerate(field_vals, 13):
        set_cell(ws, r, col, v, bg=C_NOTE, font_size=8, wrap=True, h_align="center")
    ws.row_dimensions[r].height = 30
    r += 1

    for col, rl in enumerate(field_rules, 1):
        set_cell(ws, r, col, rl, bg=C_SPECIAL, font_size=8, wrap=True, h_align="center")
    set_cell(ws, r, 12, "", bg=C_SPECIAL)
    for col, sp in enumerate(field_special, 13):
        set_cell(ws, r, col, sp, bg=C_NOTE, font_size=8, wrap=True, h_align="center")
    ws.row_dimensions[r].height = 30
    r += 1

    # --- Sample data rows ---
    def data_row(ws, r, fields_11, note_text="", bg=None):
        for col, val in enumerate(fields_11, 1):
            set_cell(ws, r, col, val, bg=bg)
        set_cell(ws, r, 12, "", bg=bg)
        ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=16)
        set_cell(ws, r, 13, note_text, bg=C_NOTE, font_size=9, italic=True, wrap=True)
        ws.row_dimensions[r].height = 22

    # Normal pin
    data_row(ws, r,
             ["1", "", "103", "SCL", "-", "L", "T", "-", "-", "-", "-"],
             "正常 pin。PKG_LOC 工具自動填，DIE_LOC 手動填（PKG_TOP_LEFT_PIN=1時）。", None)
    r += 1
    data_row(ws, r,
             ["2", "VDD33_ANA", "1", "PAD_RX_VDDA33", "-", "L", "L", "P", "", "", "RX"],
             "電源 pin。PKG_PIN_NAME 有值時 PKG PDF 使用此名稱。DIRECTION=P→紅wire。",
             C_REQUIRED)
    r += 1
    data_row(ws, r,
             ["3", "", "2", "VDD11%C%U_TOP/U_VDD11_APR0", "-", "L", "L", "P", "", "", ""],
             "含%的pin名：PKG顯示「VDD11」，APR顯示「U_VDD11_APR0」。", None)
    r += 1

    # Shared pad (PKG_NUM=0)
    data_row(ws, r,
             ["0", "", "5", "VSS_ESD", "-", "-", "L", "G", "", "", "RX"],
             "共用 die pad（無獨立 pkg pin）。同 DIE_NUM 可出現多次，每次=1條 wire。",
             C_OPTIONAL)
    r += 1
    data_row(ws, r,
             ["0", "", "5", "VSS_ESD", "-", "-", "L", "G", "", "", "RX"],
             "↑ 同一 DIE_NUM=5 再出現一次 → Combined PDF 共繪製 2 條接地 wire。",
             C_OPTIONAL)
    r += 1

    # NC
    data_row(ws, r,
             ["10", "", "0", "NC", "-", "R", "R", "-", "-", "-", ""],
             "NC pin。PKG_NUM 正整數（計入 pin 數），DIE_NUM=0，DIE_PIN_NAME=NC。",
             C_NC)
    r += 1

    # DOWNBOND
    data_row(ws, r,
             ["11", "", "0", "DOWNBOND", "-", "B", "-", "G", "-", "-", ""],
             "DOWNBOND。PKG PDF 藍色方塊；Combined PDF 繪製接地符號。DIE_NUM=0，DIRECTION=G。",
             C_DOWNBOND)
    r += 1

    # POWERCUT
    data_row(ws, r,
             ["0", "", "20", "POWERCUT_VDD", "-", "-", "T", "-", "", "", ""],
             "POWERCUT。PKG_NUM=0，DIE_PIN_NAME 含 POWERCUT。APR PDF 黑色方塊。",
             C_POWERCUT)
    r += 1

    # Inner Bond (no parens)
    data_row(ws, r,
             ["D1.94", "", "33", "VDD11%C%U_VDD11_APR0", "-", "-", "B", "P", "", "", ""],
             "Inner Bond（無括號）：D1.94 + DIE_NUM=33 → 從 die pad 94 畫線到 die pad 33。",
             C_INNERBOND)
    r += 1

    # Inner Bond (with parens)
    data_row(ws, r,
             ["(D1.94)", "", "1", "VDD11%C%U_VDD11_APR5", "-", "-", "L", "P", "", "", ""],
             "Inner Bond（有括號）：(D1.94) + DIE_NUM=1 → 從 die pad 1 畫線到 die pad 94（反向）。",
             C_INNERBOND)
    r += 1

    # More normal pins (placeholder rows)
    for i in range(3):
        pkg = 12 + i
        data_row(ws, r,
                 [str(pkg), "", str(104 + i), f"SIGNAL_{pkg}", "-",
                  "B" if pkg <= 24 else "R", "L", "-", "", "", ""],
                 "（繼續填入您的 pin 資料）", C_ALT if i % 2 == 0 else None)
        r += 1

    r += 1
    # Inner_bond template section
    section_row(ws, r, "  ▼ 區塊三：Inner_bond 模板（選填，位於檔案尾部）", ncols=16, bg="5B9BD5")
    r += 1
    ib_notes = [
        "Inner_bond 模板列不參與 pin 計數和 PDF 繪製，原樣保留到 .new 輸出，用於記錄設計意圖。"
    ]
    for n in ib_notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
        set_cell(ws, r, 1, n, italic=True, bg=C_SPECIAL)
        ws.row_dimensions[r].height = 20
        r += 1

    ib_rows = [
        ["Inner_bond", "", "1", "I0 (X.Y)", "I1 (X.Y)", "", "", "", "", "", ""],
        ["Inner_bond", "", "2", "I2 (X.Y)", "(X.Y)", "", "", "", "", "", ""],
        ["Inner_bond", "", "3", "(X.Y)", "(X.Y)", "", "", "", "", "", ""],
    ]
    for ib in ib_rows:
        data_row(ws, r, ib,
                 "Inner_bond 模板列（PKG_NUM=Inner_bond，DIE_NUM=序號，X.Y 為座標佔位符）",
                 C_SPECIAL)
        r += 1

    ws.freeze_panes = "A11"  # freeze above data rows
    return ws

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root  = os.path.dirname(script_dir)
    out_path   = os.path.join(repo_root, "docs", "ft_pad_assign_ug.xlsx")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = openpyxl.Workbook()
    build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb)
    build_sheet4(wb)
    build_sheet5(wb)

    wb.save(out_path)
    print(f"Generated: {out_path}")
    print("Sheets: 使用說明 / Header欄位 / 資料欄位 / 特殊列類型 / Pin List模板")

if __name__ == "__main__":
    main()
