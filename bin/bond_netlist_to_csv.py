#!/usr/bin/env python3
"""
bond_netlist_to_csv.py - Convert bond netlist Excel to DIE2/DIE3 overlay CSV.

Reads a bond netlist .xlsx file and extracts D2/D3 pad entries,
recalculates X,Y coordinates relative to each die's center (0,0),
and outputs CSV files in ft_pad_assign overlay format.

Usage:
    python3 bin/bond_netlist_to_csv.py docs/bond_netlist.xlsx [-o output_dir]
    python3 bin/bond_netlist_to_csv.py docs/bond_netlist.xlsx --die2-name Flash_4MB -o examples/
"""

import argparse
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip3 install openpyxl")
    sys.exit(1)


def parse_bond_netlist(xlsx_path):
    """Parse bond netlist Excel, return dict of {die_prefix: [pad_entries]}
    and d1_num_to_name mapping."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Column mapping (0-indexed)
    COL_PAD_NO = 0      # A: D1.1, D2.1, etc.
    COL_PAD_NAME = 1    # B: pad name
    COL_X = 2           # C: X coordinate
    COL_Y = 3           # D: Y coordinate
    COL_X_LEN = 4       # E: pad X length
    COL_Y_LEN = 5       # F: pad Y length
    COL_FINGER = 6      # G: finger number
    COL_UPD_NAME = 7    # H: update pad name
    COL_EDIT_NO = 8     # I: edit finger (pad) no - D1 connection
    COL_EDIT_NAME = 9   # J: edit finger name
    COL_PIN_NAME = 10   # K: pin name
    COL_PASS = 11       # L: pass flag

    dies = {}  # {prefix: [entries]}
    # D1 pad number → pin_name mapping (first occurrence wins)
    d1_num_to_name = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        pad_no = str(row[COL_PAD_NO].value).strip() if row[COL_PAD_NO].value else ""
        if not pad_no:
            continue

        # Extract die prefix (D1, D2, D3, ...)
        m = re.match(r'^(D\d+)', pad_no)
        if not m:
            continue
        prefix = m.group(1)

        pad_name = str(row[COL_PAD_NAME].value).strip() if row[COL_PAD_NAME].value else ""
        x = row[COL_X].value
        y = row[COL_Y].value
        if x is None or y is None:
            continue

        x, y = float(x), float(y)

        # D1 pad connection from EDIT_FINGER_NO (column I)
        edit_no = row[COL_EDIT_NO].value
        edit_no_str = str(edit_no).strip() if edit_no is not None else ""

        # EDIT_FINGER_NAME (column J) - power/ground indicator
        edit_name = str(row[COL_EDIT_NAME].value).strip() if row[COL_EDIT_NAME].value else ""

        # PIN_NAME (column K) - signal name
        pin_name = str(row[COL_PIN_NAME].value).strip() if row[COL_PIN_NAME].value else ""

        entry = {
            'pad_no': pad_no,
            'pad_name': pad_name,
            'x': x,
            'y': y,
            'edit_no': edit_no_str,
            'edit_name': edit_name,
            'pin_name': pin_name,
        }

        if prefix not in dies:
            dies[prefix] = []
        dies[prefix].append(entry)

        # Build D1 number → pad_name mapping for resolving D1 references
        # Use PAD_NAME (col B, e.g. QSPI0_WP) which matches DIE_PIN_NAME in pin list
        if prefix == 'D1' and pad_name and pad_name.upper() not in ('NC', 'DOWNBOND', ''):
            pad_num_match = re.match(r'^D1\.(\d+)$', pad_no)
            if pad_num_match and pad_no not in d1_num_to_name:
                d1_num_to_name[pad_no] = pad_name

    return dies, d1_num_to_name


def classify_pad(entry):
    """Classify pad as power/ground or signal based on naming conventions."""
    name_upper = entry['pad_name'].upper()
    edit_upper = entry['edit_name'].upper()
    pin_upper = entry['pin_name'].upper()

    # Power/ground indicators
    power_keywords = ['VDD', 'VCC', 'VSS', 'GND', 'VDDQ', 'VSSQ']
    for kw in power_keywords:
        if kw in name_upper or kw in edit_upper or kw in pin_upper:
            return 'power'

    return 'signal'


def extract_d1_pad_ref(entry, d1_num_to_name):
    """Extract D1 pad reference from EDIT_FINGER_NO, resolve to pad name.

    Examples:
        '(D1.74)' -> resolved pad name (e.g., 'QSPI0_WP')
        'D1.74'   -> resolved pad name
        'NC'      -> '' (no connection)
        '40'      -> '' (finger number, not pad ref)
    """
    edit_no = entry['edit_no']
    if not edit_no:
        return ""

    # Check for D1 pad reference pattern
    m = re.match(r'^\(?D1\.(\d+)\)?$', edit_no)
    if m:
        d1_key = f"D1.{m.group(1)}"
        # Resolve to pad name via D1 mapping
        return d1_num_to_name.get(d1_key, d1_key)

    return ""


def die_entries_to_csv(entries, die_name, die_prefix, d1_num_to_name):
    """Convert die entries to CSV format with coordinates relative to die center."""
    if not entries:
        return None

    # Compute bounding box center
    xs = [e['x'] for e in entries]
    ys = [e['y'] for e in entries]
    cx = (min(xs) + max(xs)) / 2.0
    cy = (min(ys) + max(ys)) / 2.0
    w = max(xs) - min(xs)
    h = max(ys) - min(ys)

    # Round die size to integers (um)
    die_w = round(w)
    die_h = round(h)

    # Build CSV rows
    num_prefix = f"{die_prefix}"  # D2 or D3
    rows = []
    for i, e in enumerate(entries, 1):
        rel_x = round(e['x'] - cx, 3)
        rel_y = round(e['y'] - cy, 3)

        d1_ref = extract_d1_pad_ref(e, d1_num_to_name)
        pad_type = classify_pad(e)

        rows.append({
            'num': f"{num_prefix}.{i}",
            'pad_name': e['pad_name'],
            'x': rel_x,
            'y': rel_y,
            'd1_pad': d1_ref,
            'type': pad_type,
        })

    return {
        'name': die_name,
        'size': f"{die_w}x{die_h}",
        'rows': rows,
        'center': (cx, cy),
        'bbox': (min(xs), min(ys), max(xs), max(ys)),
    }


def write_die_csv(csv_data, die_prefix, output_path):
    """Write DIE overlay CSV file."""
    # die_prefix is 'D2' or 'D3'; parser expects DIE2_NAME / DIE3_NAME
    die_num = die_prefix[1]  # '2' or '3'
    name_key = f"DIE{die_num}_NAME"
    size_key = "DIE_SIZE"
    loc_key = f"DIE{die_num}_LOC"
    num_key = f"{die_prefix}_NUM"  # D2_NUM or D3_NUM

    lines = []
    lines.append(f"{name_key} : {csv_data['name']},,,,,")
    lines.append(f"{size_key} : {csv_data['size']},,,,,")
    lines.append(f"{loc_key} : 0,0,,,,")  # placeholder, user adjusts
    lines.append("PLACEMENT : R0,,,,,")
    lines.append(",,,,,")
    lines.append(f"{num_key},{die_prefix}_PAD_NAME,X,Y,D1_PAD,TYPE")

    for r in csv_data['rows']:
        d1 = r['d1_pad'] if r['d1_pad'] else ''
        typ = r['type']
        lines.append(f"{r['num']},{r['pad_name']},{r['x']},{r['y']},{d1},{typ}")

    with open(output_path, 'w') as f:
        f.write('\n'.join(lines) + '\n')

    return len(csv_data['rows'])


def main():
    parser = argparse.ArgumentParser(
        description='Convert bond netlist Excel to DIE2/DIE3 overlay CSV')
    parser.add_argument('input', help='Bond netlist .xlsx file')
    parser.add_argument('-o', '--output-dir', default='.', help='Output directory')
    parser.add_argument('--die2-name', default=None,
                        help='DIE2 display name (default: from filename)')
    parser.add_argument('--die3-name', default=None,
                        help='DIE3 display name (default: from filename)')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: File not found: {args.input}")
        sys.exit(1)

    # Default die name from filename
    base = os.path.splitext(os.path.basename(args.input))[0]
    # Strip common suffixes like .bond_netlist-241203-180841
    base_clean = re.sub(r'\.bond_netlist.*$', '', base)
    default_name = base_clean if base_clean else base

    die2_name = args.die2_name or default_name
    die3_name = args.die3_name or default_name

    os.makedirs(args.output_dir, exist_ok=True)

    print(f"Reading: {args.input}")
    dies, d1_num_to_name = parse_bond_netlist(args.input)

    if not dies:
        print("ERROR: No die entries found in input file")
        sys.exit(1)

    print(f"Found dies: {', '.join(sorted(dies.keys()))}")

    generated = []

    # Process D2
    if 'D2' in dies:
        csv_data = die_entries_to_csv(dies['D2'], die2_name, 'D2', d1_num_to_name)
        if csv_data:
            out_path = os.path.join(args.output_dir, 'DIE2_bond.csv')
            n = write_die_csv(csv_data, 'D2', out_path)
            cx, cy = csv_data['center']
            print(f"\nDIE2: {n} pads, center=({cx:.1f},{cy:.1f}), "
                  f"size={csv_data['size']} um")
            print(f"  BBox: ({csv_data['bbox'][0]:.1f},{csv_data['bbox'][1]:.1f})"
                  f" - ({csv_data['bbox'][2]:.1f},{csv_data['bbox'][3]:.1f})")
            print(f"  Output: {out_path}")
            generated.append('DIE2')

    # Process D3
    if 'D3' in dies:
        csv_data = die_entries_to_csv(dies['D3'], die3_name, 'D3', d1_num_to_name)
        if csv_data:
            out_path = os.path.join(args.output_dir, 'DIE3_bond.csv')
            n = write_die_csv(csv_data, 'D3', out_path)
            cx, cy = csv_data['center']
            print(f"\nDIE3: {n} pads, center=({cx:.1f},{cy:.1f}), "
                  f"size={csv_data['size']} um")
            print(f"  BBox: ({csv_data['bbox'][0]:.1f},{csv_data['bbox'][1]:.1f})"
                  f" - ({csv_data['bbox'][2]:.1f},{csv_data['bbox'][3]:.1f})")
            print(f"  Output: {out_path}")
            generated.append('DIE3')

    if not generated:
        print("WARNING: No D2 or D3 entries found. Only D1 present.")
        sys.exit(0)

    # Summary
    print(f"\n--- Summary ---")
    print(f"Generated: {', '.join(generated)}")
    print(f"\nCoordinate transform: pad_XY - die_center = relative_XY")
    print(f"Die center computed from bounding box of all pads.")
    print(f"\nNOTE: DIE2_LOC / DIE3_LOC set to '0,0' placeholder.")
    print(f"      Adjust to actual offset from DIE1 origin in um.")


if __name__ == '__main__':
    main()
